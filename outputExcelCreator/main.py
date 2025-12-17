import openpyxl as op
import configparser as cp

import os
import sys

from unidecode import unidecode as ud
from rapidfuzz import process, fuzz

import auxiliar as aux
import Global as glb

"""
Paso 1: Crear una matriz de productos. Cada fila dispone de las siguientes columnas:

* Nombre del producto
* Categoría
* Ratio
* Nombre del formato
* Alérgenos
* Precio principal
* Precio añadido
* Ruta de la imagen dentro del directorio de imágenes de Ágora definido en setting.ini
* Tipo de preparación
* Orden de preparación
* Texto de documento/comanda
* Texto de botón
* Descripción para la carta digital
"""

productos_salida = []
nombres_productos_salida = []

# Creación del archivo de LOG

with open("temp.txt", "r") as f:
    lineas = f.readlines()

project = lineas[0][:-1]
log_file_name = lineas[1]

if not os.path.exists("projects/" + project + "/log_files"):
    os.makedirs("projects/" + project + "/log_files")

glb.log_file = open("projects/" + project + "/log_files/%s" % log_file_name + ".log", "a")
glb.log_file.truncate(0)

"""
Paso 2: Leer, del fichero de entrada, los productos definidos en la hoja "Productos".
Los datos que estén incompletos serán consultados en la cache.
Si dichos datos no son localizados, se inferirán o se consultarán al usuario.
"""

# Se leen, del fichero settings.ini, los nombres de los ficheros Excel de entrada y salida,
# así como de la carpeta de imágenes de Ágora
glb.datos_import = "projects/" + project + "/entrada.xlsx"
glb.datos_export = "projects/" + project + "/salida.xlsx"
glb.dir_imagenes = "imagenesAgora"

glb.ruta_cache = "outputExcelCreator\\data\\cache.xlsx"

libro = op.load_workbook(glb.datos_import)
hoja_productos = libro["Productos"]

productos_padre = []

limit = 0

while hoja_productos["A" + str(limit + 2)].value is not None:
    limit += 1

productos_en_cache = []
archivo_cache = op.load_workbook(glb.ruta_cache)
hoja_productos_cache = archivo_cache["Productos"]

c = 2
while hoja_productos_cache["A" + str(c)].value is not None:
    producto = str(hoja_productos_cache["A" + str(c)].value)
    if producto not in productos_en_cache:
        productos_en_cache.append(producto)

    c += 1

i = 2
j = 2
while hoja_productos["A" + str(i)].value is not None:
    nombre = str(hoja_productos["A" + str(i)].value)

    # Buscar en la cache y actualizar, si procede...
    if nombre in productos_en_cache:
        while hoja_productos_cache["A" + str(j)].value is not None:
            if nombre == str(hoja_productos_cache["A" + str(j)].value):
                # Actualizar familia en la cache
                if hoja_productos["C" + str(i)].value is not None and hoja_productos_cache["B" + str(j)].value != \
                        hoja_productos["C" + str(i)].value:
                    hoja_productos_cache["B" + str(j)].value = hoja_productos["C" + str(i)].value
                    glb.log_file.write("ACTUALIZADA EN CACHE LA FAMILIA DE %s AL VALOR %s\n" % (
                    nombre, hoja_productos["C" + str(i)].value))

                # Actualizar tipo de preparación en la cache
                if hoja_productos["D" + str(i)].value is not None and hoja_productos_cache["C" + str(j)].value != \
                        hoja_productos["D" + str(i)].value:
                    hoja_productos_cache["C" + str(j)].value = hoja_productos["D" + str(i)].value
                    glb.log_file.write("ACTUALIZADO EN CACHE EL TIPO DE PREPARACIÓN DE %s AL VALOR %s\n" % (
                    nombre, hoja_productos["D" + str(i)].value))

                # Actualizar orden de preparación en la cache
                if hoja_productos["E" + str(i)].value is not None and hoja_productos_cache["D" + str(j)].value != \
                        hoja_productos["E" + str(i)].value:
                    hoja_productos_cache["D" + str(j)].value = hoja_productos["E" + str(i)].value
                    glb.log_file.write("ACTUALIZADO EN CACHE EL ORDEN DE PREPARACIÓN DE %s AL VALOR %s\n" % (
                    nombre, hoja_productos["E" + str(i)].value))

                # Actualizar ruta de la imagen en la cache
                if hoja_productos["F" + str(i)].value is not None and hoja_productos_cache["E" + str(j)].value != \
                        hoja_productos["F" + str(i)].value:
                    hoja_productos_cache["E" + str(j)].value = hoja_productos["F" + str(i)].value
                    glb.log_file.write("ACTUALIZADA EN CACHE LA RUTA DE IMAGEN DE %s AL VALOR %s\n" % (
                    nombre, hoja_productos["F" + str(i)].value))

                # Actualizar descripción en la cache
                if hoja_productos["I" + str(i)].value is not None and hoja_productos_cache["I" + str(j)].value != \
                        hoja_productos["I" + str(i)].value:
                    hoja_productos_cache["I" + str(j)].value = hoja_productos["I" + str(i)].value
                    glb.log_file.write("ACTUALIZADA EN CACHE LA DESCRIPCIÓN DE %s AL VALOR %s\n" % (
                    nombre, hoja_productos["I" + str(i)].value))

                # Actualizar existencia de imagen en la cache
                if hoja_productos_cache["E" + str(j)].value is not None and os.path.exists(
                        hoja_productos_cache["E" + str(j)].value):
                    hoja_productos_cache["K" + str(j)].value = "Sí"
                    glb.log_file.write("EXISTE LA IMAGEN DEL PRODUCTO %s\n" % nombre)

                    # Actualizar parentesco de la imagen y el nombre del producto
                    match = process.extractOne(
                        aux.normalizar_nombre_archivo(nombre),
                        glb.dir_imagenes,
                        scorer=fuzz.ratio,
                        score_cutoff=0.9)

                    if match:
                        _, score, _ = match
                        parentesco_score = round(score / 100, 2)
                        hoja_productos_cache["J" + str(j)].value = str(parentesco_score)
                        glb.log_file.write("EL PARENTESCO ENTRE EL PRODUCTO %s Y LA RUTA DE SU IMAGEN ES %.2f\n" % (nombre, parentesco_score))

                else:
                    hoja_productos_cache["K" + str(j)].value = "No"
                    glb.log_file.write("NO EXISTE LA IMAGEN DEL PRODUCTO %s\n" % nombre)

            j += 1
    else:
        # Si no está, se añade el producto en la cache con los datos que ya se conocen

        # Actualizar nombre en la cache
        if hoja_productos["A" + str(i)].value is not None:
            hoja_productos_cache["A" + str(c)].value = str(hoja_productos["A" + str(i)].value)
        else:
            hoja_productos_cache["A" + str(c)].value = ""

        # Actualizar categoría en la cache
        if hoja_productos["C" + str(i)].value is not None:
            hoja_productos_cache["B" + str(c)].value = hoja_productos["C" + str(i)].value
        else:
            hoja_productos_cache["B" + str(c)].value = ""

        # Actualizar tipo de preparación en la cache
        if hoja_productos["D" + str(i)].value is not None:
            hoja_productos_cache["C" + str(c)].value = hoja_productos["D" + str(i)].value
        else:
            hoja_productos_cache["C" + str(c)].value = ""

        # Actualizar orden de preparación en la cache
        if hoja_productos["E" + str(i)].value is not None:
            hoja_productos_cache["D" + str(c)].value = hoja_productos["E" + str(i)].value
        else:
            hoja_productos_cache["D" + str(c)].value = ""

        # Actualizar ruta de la imagen en la cache
        if hoja_productos["F" + str(i)].value is not None:
            hoja_productos_cache["E" + str(c)].value = hoja_productos["F" + str(i)].value
        else:
            hoja_productos_cache["E" + str(c)].value = ""

        # Actualizar descripción en la cache
        if hoja_productos["I" + str(i)].value is not None:
            hoja_productos_cache["I" + str(c)].value = hoja_productos["I" + str(i)].value
        else:
            hoja_productos_cache["I" + str(c)].value = ""

        # Actualizar descripción en la cache
        if hoja_productos["I" + str(i)].value is not None and hoja_productos_cache["I" + str(j)].value != \
                hoja_productos["I" + str(i)].value:
            hoja_productos_cache["I" + str(j)].value = hoja_productos["I" + str(i)].value
            glb.log_file.write("ACTUALIZADA EN CACHE LA DESCRIPCIÓN DE %s AL VALOR %s\n" % (
            nombre, hoja_productos["I" + str(i)].value))

        # Actualizar existencia de imagen en la cache
        if hoja_productos_cache["E" + str(j)].value is not None and os.path.exists(
                hoja_productos_cache["E" + str(j)].value):
            hoja_productos_cache["K" + str(j)].value = "Sí"
            glb.log_file.write("EXISTE LA IMAGEN DEL PRODUCTO %s\n" % nombre)

            # Actualizar parentesco de la imagen y el nombre del producto
            match = process.extractOne(
                aux.normalizar_nombre_archivo(nombre),
                glb.dir_imagenes,
                scorer=fuzz.ratio,
                score_cutoff=0.9)

            if match:
                _, score, _ = match
                parentesco_score = round(score / 100, 2)
                hoja_productos_cache["J" + str(j)].value = str(parentesco_score)
                glb.log_file.write("EL PARENTESCO ENTRE EL PRODUCTO %s Y LA RUTA DE SU IMAGEN ES %.2f\n" % (nombre, parentesco_score))

        else:
            hoja_productos_cache["K" + str(j)].value = "No"
            glb.log_file.write("NO EXISTE LA IMAGEN DEL PRODUCTO %s\n" % nombre)

        glb.log_file.write("INSERTADO EN CACHE EL PRODUCTO %s, FILA %s, VALORES: %s, %s, %s, %s, %s\n" % (
            str(hoja_productos["A" + str(i)].value),
            str(c),
            hoja_productos["C" + str(i)].value,
            hoja_productos["D" + str(i)].value,
            hoja_productos["E" + str(i)].value,
            hoja_productos["F" + str(i)].value,
            hoja_productos["I" + str(i)].value
        ))

        c += 1

    categoria = hoja_productos["C" + str(i)].value

    # Si el producto no tiene una categoría informada, se consultará en la cache
    if categoria is None:
        categoria = aux.find_in_cache("CAT", nombre)
        hoja_productos["C" + str(i)].value = categoria

    ratio = hoja_productos["B" + str(i)].value
    base = nombre
    formato = ""
    producto_con_formato = base

    if ratio is not None:
        base = nombre.split("(")[0][:-1]
        formato = nombre.split("(")[1][:-1]

        producto_con_formato = formato + " " + base

    alergenos = aux.get_allergens(nombre, categoria)
    precio_principal = hoja_productos["G" + str(i)].value
    precio_addin = hoja_productos["H" + str(i)].value
    ruta_imagen = hoja_productos["F" + str(i)].value
    ruta_imagen_padre = ""

    # Si el producto no tiene una ruta de imagen informada, se consultará en la cache
    if ratio is None:
        if ruta_imagen is None or not os.path.exists(os.getcwd() + "\\" + ruta_imagen):
            ruta_imagen = aux.find_in_cache("IMG", nombre=base, categoria=categoria)
    else:
        ruta_imagen_padre = aux.find_in_cache("IMG", nombre=base, categoria=categoria)
        if ruta_imagen is None or not os.path.exists(os.getcwd() + "\\" + ruta_imagen):
            ruta_imagen \
                = glb.dir_imagenes + "\\" + ud(categoria.lower().split(" ")[0]) + "\\" + ud(formato.lower()) + ".jpg"

    hoja_productos["F" + str(i)].value = ruta_imagen

    abierto = True
    while abierto:
        try:
            op.load_workbook(glb.datos_import)
            abierto = False
        except IOError:
            abierto = True

    libro.save(glb.datos_import)

    tipo_preparacion = hoja_productos["D" + str(i)].value

    # Si el producto no tiene un tipo de preparacion informada, se consultará en la cache
    if tipo_preparacion is None:
        tipo_preparacion = aux.find_in_cache("TIP", nombre=nombre, categoria=categoria)
        hoja_productos["D" + str(i)].value = tipo_preparacion

    orden_preparacion = hoja_productos["E" + str(i)].value

    # Si el producto no tiene un orden de preparacion informada, se consultará en la cache
    if orden_preparacion is None:
        orden_preparacion = aux.find_in_cache("ORD", nombre=nombre, categoria=categoria)
        hoja_productos["E" + str(i)].value = orden_preparacion

    texto_boton = ""
    texto_ticket = ""
    texto_boton_padre = ""
    texto_ticket_padre = ""
    texto_boton_vacio = True
    texto_ticket_vacio = True

    # Si no se encuentran los datos en la cache, el programa los generará
    j = 2
    while hoja_productos_cache["A" + str(j)].value is not None:
        if hoja_productos_cache["A" + str(j)].value == nombre:
            if hoja_productos_cache["F" + str(j)].value is not None:
                texto_boton = str(hoja_productos_cache["F" + str(j)].value)
                texto_boton_vacio = False

            if hoja_productos_cache["G" + str(j)].value is not None:
                texto_ticket = str(hoja_productos_cache["G" + str(j)].value)
                texto_ticket_vacio = False

            break

        j += 1

    # if texto_boton == "" or texto_ticket == "":
    if texto_boton == "" or texto_ticket == "":
        texto_boton, texto_ticket = aux.text_generator(base, ruta_imagen != "")

    if len(texto_boton) > 10 and ruta_imagen == "":
        ruta_imagen = ""

    if ratio is not None:
        # Buscar los textos del padre en la cache
        k = 2
        texto_boton_padre = ""
        texto_ticket_padre = ""

        while hoja_productos_cache["A" + str(k)].value is not None:
            if hoja_productos_cache["A" + str(k)].value == base:
                if hoja_productos_cache["F" + str(k)].value is not None:
                    texto_boton_padre = str(hoja_productos_cache["F" + str(k)].value)

                if hoja_productos_cache["G" + str(k)].value is not None:
                    texto_ticket_padre = str(hoja_productos_cache["G" + str(k)].value)

                break

            k += 1

        if texto_boton_padre == "":
            texto_boton_padre = texto_boton

        if texto_ticket_padre == "":
            texto_ticket_padre = texto_ticket

        # if texto_boton == "" or texto_ticket == "":
        # Si no se encuentran los datos en la cache, el programa los generará
        texto_boton = formato
        abreviado = aux.get_abreviated_format(formato)
        _, texto_ticket = aux.text_generator(base, ruta_imagen != "", 19 - len(abreviado))

        texto_ticket = abreviado + " " + texto_ticket

    if texto_boton_vacio:
        hoja_productos_cache["F" + str(j)].value = texto_boton
        glb.log_file.write("ACTUALIZADO EN CACHE EL TEXTO DE BOTÓN DE %s AL VALOR %s\n" % (nombre, texto_boton))

    if texto_ticket_vacio:
        hoja_productos_cache["G" + str(j)].value = texto_ticket
        hoja_productos_cache["H" + str(j)].value = texto_ticket
        glb.log_file.write("ACTUALIZADO EN CACHE EL TEXTO DE TICKET DE %s AL VALOR %s\n" % (nombre, texto_ticket))

    descripcion = hoja_productos["I" + str(i)].value

    if ratio is not None:
        # Añadir el producto padre si aún no se había hecho
        if ruta_imagen_padre is None or not os.path.exists(os.getcwd() + "\\" + ruta_imagen_padre):
            ruta_imagen_padre = glb.dir_imagenes + "\\" + ud(categoria.lower().split(" ")[0]) + "\\" + ud(
                base.lower()) + ".jpg"
            ruta_imagen_padre = ruta_imagen_padre.replace(" ", "-")
            
        producto_padre = [
            base,
            categoria,
            None,
            None,
            alergenos,
            None,
            None,
            os.getcwd() + "\\" + ruta_imagen_padre if ruta_imagen_padre is not None and ruta_imagen_padre != "" else None,
            tipo_preparacion,
            orden_preparacion,
            texto_boton_padre,
            texto_ticket_padre,
            descripcion
        ]

        # Añadir el producto padre a la cache si no estaba

        j = 2
        padre_en_cache = False
        while hoja_productos_cache["A" + str(j)].value is not None:
            if hoja_productos_cache["A" + str(j)].value == base:
                padre_en_cache = True
                break

            j += 1

        if not padre_en_cache:
            hoja_productos_cache["A" + str(j)].value = base
            hoja_productos_cache["B" + str(j)].value = categoria
            hoja_productos_cache["C" + str(j)].value = tipo_preparacion
            hoja_productos_cache["D" + str(j)].value = orden_preparacion
            hoja_productos_cache["E" + str(j)].value = ruta_imagen_padre if ruta_imagen_padre is not None else ""
            hoja_productos_cache["F" + str(j)].value = texto_boton_padre
            hoja_productos_cache["G" + str(j)].value = texto_ticket_padre
            hoja_productos_cache["H" + str(j)].value = texto_ticket_padre
            hoja_productos_cache["I" + str(j)].value = descripcion

            productos_en_cache.append(base)
            c += 1

            if ruta_imagen_padre is not None and os.path.exists(ruta_imagen_padre):
                hoja_productos_cache["K" + str(j)].value = "Sí"

                match = process.extractOne(
                    aux.normalizar_nombre_archivo(base),
                    glb.dir_imagenes,
                    scorer=fuzz.ratio,
                    score_cutoff=0.9)

                if match:
                    _, score, _ = match
                    parentesco_score = round(score / 100, 2)
                    hoja_productos_cache["J" + str(j)].value = str(parentesco_score)

            else:
                hoja_productos_cache["K" + str(j)].value = "No"

            glb.log_file.write("INSERTADO EN CACHE EL PRODUCTO %s, FILA %s, VALORES: %s, %s, %s, %s, %s\n" % (
                str(hoja_productos["A" + str(j)].value),
                str(j),
                hoja_productos["C" + str(j)].value,
                hoja_productos["D" + str(j)].value,
                hoja_productos["E" + str(j)].value,
                hoja_productos["F" + str(j)].value,
                hoja_productos["I" + str(j)].value
            ))

        if base not in nombres_productos_salida:
            productos_salida.append(producto_padre)
            nombres_productos_salida.append(base)

        producto_hijo = [
            base,
            categoria,
            ratio,
            producto_con_formato,
            None,
            precio_principal,
            precio_addin,
            os.getcwd() + "\\" + ruta_imagen if ruta_imagen is not None and ruta_imagen != "" else None,
            None,
            None,
            texto_boton,
            texto_ticket,
            None
        ]
        productos_salida.append(producto_hijo)
        nombres_productos_salida.append(producto_con_formato)

    else:
        producto_sin_formato = [
            base,
            categoria,
            ratio,
            None,
            alergenos,
            precio_principal,
            precio_addin,
            os.getcwd() + "\\" + ruta_imagen if ruta_imagen is not None and ruta_imagen != "" else None,
            tipo_preparacion,
            orden_preparacion,
            texto_boton,
            texto_ticket,
            descripcion
        ]
        productos_salida.append(producto_sin_formato)
        nombres_productos_salida.append(base)

    i += 1

    sys.stdout.write("\r%d / %d productos procesados" % (i - 2, limit))
    sys.stdout.flush()

# Se guardan los posibles cambios que puedan haberse dado
abierto = True
while abierto:
    try:
        op.load_workbook(glb.datos_import)
        abierto = False
    except IOError:
        abierto = True

libro.save(glb.datos_import)

productos_salida = sorted(productos_salida, key=lambda x: int(x[2]) if x[2] is not None else 0)  # Ordena por formato
productos_salida = sorted(productos_salida, key=lambda x: x[0])  # Ordena por nombre
productos_salida = sorted(productos_salida, key=lambda x: x[1])  # Ordena por categoría

"""
Paso 3: Una vez obtenida la matriz de productos, estos se guardan en la hoja "Productos" del Excel de salida. Algunos
datos, como el condicional de vendible como principal o añadido, dependen de información ya indicada en la matriz
"""

salida = op.load_workbook("outPutExcelCreator\\data\\plantilla_vacia.xlsx")
hoja_productos_salida = salida["Productos"]

i = 2
for producto in productos_salida:
    hoja_productos_salida["A" + str(i)] = producto[1]
    hoja_productos_salida["B" + str(i)] = producto[1]
    hoja_productos_salida["C" + str(i)] = producto[4]
    hoja_productos_salida["E" + str(i)] = producto[0]
    hoja_productos_salida["F" + str(i)] = producto[3]
    hoja_productos_salida["G" + str(i)] = producto[2]
    hoja_productos_salida["K" + str(i)] = "7" if producto[3] is None else None
    hoja_productos_salida["P" + str(i)] = "No" if producto[5] is None else "Sí"
    hoja_productos_salida["Q" + str(i)] = "No" if producto[6] is None else "Sí"
    hoja_productos_salida["S" + str(i)] = "0xFFBACDE2"
    hoja_productos_salida["T" + str(i)] = producto[10]
    hoja_productos_salida["U" + str(i)] = producto[7] if producto[7] is not None and os.path.exists(
        producto[7]) else None
    hoja_productos_salida["V" + str(i)] = producto[8] if producto[3] is None else None
    hoja_productos_salida["W" + str(i)] = producto[9] if producto[3] is None else None
    hoja_productos_salida["X" + str(i)] = producto[11]
    hoja_productos_salida["Y" + str(i)] = producto[11]
    hoja_productos_salida["Z" + str(i)] = producto[12]
    hoja_productos_salida["AB" + str(i)] = producto[5]
    hoja_productos_salida["AC" + str(i)] = producto[6]

    i += 1

"""
Paso 4: Crear una matriz de notas de preparación. Cada fila dispone de las siguientes columnas:

* Texto de la nota
* Categorías afectadas por la nota de preparación
* Texto de botón
* Imagen de la nota de preparación. Deberá aparecer en el directorio de imágenes de Ágora, bajo la carpeta "notas"
"""

notas = []

"""
Paso 5: Leer, del fichero de entrada, las notas de preparación definidas en la hoja "Notas".
"""

hoja_notas = libro["Notas"]

i = 2
while hoja_notas["A" + str(i)].value is not None:
    texto = hoja_notas["A" + str(i)].value
    categorias = hoja_notas["B" + str(i)].value
    _, texto_boton = aux.text_generator(texto, True, 10)
    ruta_imagen = os.getcwd() + "\\" + glb.dir_imagenes + "\\notas\\" + ud(texto.replace(" ", "-").lower() + ".jpg")

    nota = [
        texto,
        categorias,
        texto_boton,
        ruta_imagen
    ]

    notas.append(nota)

    i += 1

hoja_notas_salida = salida["Notas de preparación"]

i = 2
for nota in notas:
    hoja_notas_salida["A" + str(i)].value = nota[0]
    hoja_notas_salida["B" + str(i)].value = "0"
    hoja_notas_salida["D" + str(i)].value = nota[1]
    hoja_notas_salida["E" + str(i)].value = nota[2]
    hoja_notas_salida["F" + str(i)].value = "#BACDE2"
    hoja_notas_salida["G" + str(i)].value = nota[3] if os.path.exists(nota[3]) else None

    i += 1

"""
Paso 6: Crear una matriz de usuarios. Cada fila dispone de las siguientes columnas:

* Nombre
* Perfil
* Clave de TPV
* Clave de CMD
* Clave de Administración Web
* Imagen del usuario
"""

usuarios_salida = []

"""
Paso 7: Leer, del fichero de entrada, los usuarios definidos en la hoja "Notas".
"""

hoja_usuarios = libro["Usuarios"]

# El usuario Infogral debe ser el primero de todos

pos_key = aux.generate_random_key(False)
admin_web_key = aux.generate_random_key(True)

hoja_usuarios_salida = salida["Usuarios"]

hoja_usuarios_salida["A2"].value = "Infogral"
hoja_usuarios_salida["C2"].value = pos_key
hoja_usuarios_salida["D2"].value = pos_key
hoja_usuarios_salida["E2"].value = admin_web_key
hoja_usuarios_salida["G2"].value = "Administrador"
hoja_usuarios_salida["P2"].value = "Infogral"
hoja_usuarios_salida["Q2"].value = os.getcwd() + "\\" + glb.dir_imagenes + "\\usuarios\\Usuario-Infogral.png"
hoja_usuarios_salida["R2"].value = "#FFFFFF"
hoja_usuarios_salida["S2"].value = "No"
hoja_usuarios_salida["T2"].value = "No"
hoja_usuarios_salida["U2"].value = 0
hoja_usuarios_salida["Z2"].value = "FALSO"

i = 2
while hoja_usuarios["A" + str(i)].value is not None:
    nombre = hoja_usuarios["A" + str(i)].value
    perfil = hoja_usuarios["B" + str(i)].value
    clave_tpv = hoja_usuarios["C" + str(i)].value
    clave_cmd = hoja_usuarios["D" + str(i)].value
    clave_admin = hoja_usuarios["E" + str(i)].value
    imagen = os.getcwd() + "\\" + glb.dir_imagenes + "\\usuarios\\" + perfil + ".jpg"

    clave_tpv = aux.check_password(clave_tpv, perfil, False)
    clave_cmd = aux.check_password(clave_cmd, perfil, False)
    clave_admin = aux.check_password(clave_admin, perfil, True)

    usuario = [
        nombre,
        perfil,
        clave_tpv,
        clave_cmd,
        clave_admin,
        imagen
    ]

    usuarios_salida.append(usuario)

    i += 1

i = 3
for usuario in usuarios_salida:
    hoja_usuarios_salida["A" + str(i)].value = usuario[0]
    hoja_usuarios_salida["C" + str(i)].value = usuario[2] if usuario[2] is not None else ""
    hoja_usuarios_salida["D" + str(i)].value = usuario[3] if usuario[3] is not None else ""
    hoja_usuarios_salida["E" + str(i)].value = usuario[4]
    hoja_usuarios_salida["G" + str(i)].value = usuario[1]
    hoja_usuarios_salida["P" + str(i)].value = usuario[0]
    hoja_usuarios_salida["Q" + str(i)].value = usuario[5] if os.path.exists(usuario[5]) else None
    hoja_usuarios_salida["R" + str(i)].value = "#FFFFFF"
    hoja_usuarios_salida["S" + str(i)].value = "Sí"
    hoja_usuarios_salida["T" + str(i)].value = "Sí"
    hoja_usuarios_salida["U" + str(i)].value = 0
    hoja_usuarios_salida["Z" + str(i)].value = "FALSO"

    i += 1

"""
Paso 8: Crear una matriz de centros de venta. Cada fila dispone de las siguientes columnas:

* Nombre
* Listado de ubicaciones
"""

centros_salida = []

"""
Paso 9: Leer, del fichero de entrada, los centros de venta definidos en la hoja "Centros de venta".
"""

hoja_centros = libro["Centros de venta"]

i = 2
while hoja_centros["A" + str(i)].value is not None:
    nombre = hoja_centros["A" + str(i)].value
    cantidad = int(hoja_centros["B" + str(i)].value) if hoja_centros["B" + str(i)].value is not None else 1
    nomenclatura = "" if hoja_centros["C" + str(i)].value is None else hoja_centros["C" + str(i)].value
    lista_ubicaciones = nomenclatura + "1"

    for j in range(2, cantidad + 1):
        lista_ubicaciones += ", " + nomenclatura + str(j)

    centro_salida = [
        nombre,
        lista_ubicaciones
    ]

    centros_salida.append(centro_salida)

    i += 1

hoja_centros_salida = salida["Centros de Venta"]

i = 2
for centro in centros_salida:
    hoja_centros_salida["A" + str(i)].value = centro[0]
    hoja_centros_salida["B" + str(i)].value = "General"
    hoja_centros_salida["C" + str(i)].value = "Almacén General"
    hoja_centros_salida["D" + str(i)].value = "Nunca"
    hoja_centros_salida["E" + str(i)].value = "Nunca"
    hoja_centros_salida["F" + str(i)].value = centro[0]
    hoja_centros_salida["G" + str(i)].value = "#BACDE2"
    hoja_centros_salida["I" + str(i)].value = centro[1]
    hoja_centros_salida["J" + str(i)].value = "0"
    hoja_centros_salida["K" + str(i)].value = "No"

    i += 1

"""
Paso 10: Guardar los datos generados en el Excel de salida.
"""

glb.log_file.close()
salida.save(glb.datos_export)
archivo_cache.save(glb.ruta_cache)
print()
