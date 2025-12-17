import os
import sys

import openpyxl as op
import configparser as cp
import deepl as dl

from unidecode import unidecode as ud

import Global as glb
import auxiliar as aux

"""
Paso 1: Leer la hoja de Productos del Excel de Salida y obtener los distintos nombres de categorías
"""

with open("temp.txt", "r") as f:
    lineas = f.readlines()

project = lineas[0][:-1]
log_file_name = lineas[1]

# Se leen, del fichero settings.ini, los nombres de los ficheros Excel de entrada y salida,
# así como de la carpeta de imágenes de Ágora
glb.datos_import = "projects/" + project + "/entrada.xlsx"
glb.datos_export = "projects/" + project + "/salida.xlsx"
glb.dir_imagenes = "imagenesAgora"
glb.log_file = open("projects/" + project + "/log_files/%s" % log_file_name + ".log", "a")

glb.ruta_cache = "outputExcelCreator/data/cache.xlsx"

translator = dl.Translator(glb.auth_key)
archivo_cache = op.load_workbook(glb.ruta_cache)
traducciones = archivo_cache["Traducciones"]

script_sql = open("sql/categories-and-families.sql", "w")

categorias_distintas = []

salida = op.load_workbook(glb.datos_export)

hoja_productos_salida = salida["Productos"]

limit = 0

i = 2
while hoja_productos_salida["A" + str(i)].value is not None:
    if hoja_productos_salida["A" + str(i)].value not in categorias_distintas:
        categorias_distintas.append(hoja_productos_salida["A" + str(i)].value)

    i += 1

categorias_entrada = []

entrada = op.load_workbook(glb.datos_import)

hoja_productos_entrada = entrada["Productos"]

i = 2
while hoja_productos_entrada["C" + str(i)].value is not None:
    if hoja_productos_entrada["C" + str(i)].value not in categorias_entrada:
        categorias_entrada.append(hoja_productos_entrada["C" + str(i)].value)

    i += 1

fuera_de_entrada = len(categorias_entrada)

cache_traducciones_size = 0

while traducciones["A" + str(cache_traducciones_size + 2)].value is not None:
    cache_traducciones_size += 1

"""
Paso 2: Obtener, por cada categoría, su texto abreviado e imagen
"""

categorias = []

for categoria_distinta in categorias_distintas:
    nombre = categoria_distinta
    imagen = (os.getcwd() +
              "\\" +
              glb.dir_imagenes +
              "\\categorias-carta-digital\\" +
              ud(nombre).replace(" ", "-").lower() +
              ".jpg")
    texto, _ = aux.text_generator(nombre, os.path.exists(imagen))

    categoria = [
        nombre,
        imagen,
        texto
    ]

    categorias.append(categoria)

    limit += 1

query = (
    "SET NOCOUNT ON;\n\n"
    
    "DECLARE @familia NVARCHAR(255);\n"
    "DECLARE @categoria NVARCHAR(255);\n"
    "DECLARE @idFamilia INT;\n"
    "DECLARE @idTraduccion INT;\n"
    "DECLARE @idCategoria INT;\n\n"
    "DECLARE @rutaImagen VARCHAR(255);\n"
    "DECLARE @textoBoton VARCHAR(10);\n"
    "DECLARE @SQL_QUERY NVARCHAR(MAX);\n"
    "DECLARE @contenido VARBINARY(MAX);\n"
    "DECLARE @saleableProducts INT;\n\n"
    
    "INSERT INTO [igtpos].[dbo].[IndexedGroup]\n"
    "SELECT 'ProductGroups_1'\n"
    "WHERE NOT EXISTS\n"
    "(\n"
    "\tSELECT 1\n"
    "\tFROM [igtpos].[dbo].[IndexedGroup]\n"
    "\tWHERE KeyName = 'ProductGroups_1'\n"
    ");\n\n"
)

a = 1
for categoria in categorias:
    """
    Paso 3: Insertar la familia en la tabla [igtpos].[dbo].Family y la categoría en la tabla [igtpos].[dbo].Category
    """

    nombre = categoria[0]
    ruta_imagen = categoria[1]
    texto_boton = categoria[2]

    query += (
        "SET @familia = '%s';\n"
        "SET @categoria = @familia;\n"
        "SET @rutaImagen = '%s';\n"
        "SET @textoBoton = '%s';\n"
        "SET @saleableProducts = 0;\n\n"
        % (nombre, ruta_imagen, texto_boton)
    )

    # Si no existe la familia, se crea

    query += (
        "INSERT INTO [igtpos].[dbo].Family\n"
        "SELECT NULL, '%s', 0, 0, 0, NULL, '%s', '0xFFBACDE2', 0x0, NULL, 1\n"
        "WHERE NOT EXISTS\n"
        "\t(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].Family\n"
        "\tWHERE [Name] = '%s'\n"
        ");\n\n"
        
        "INSERT INTO [igtpos].[dbo].Category\n"
        "SELECT NULL, '%s', '%s', '0xFFBACDE2', 0x0, NULL, 1, 1, 1, 1\n"
        "WHERE NOT EXISTS\n"
        "\t(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].Category\n"
        "\tWHERE [Name] = '%s'\n"
        ");\n\n"
        
        "SET @idFamilia =\n"
        "(\n"
        "\tSELECT Id\n"
        "\tFROM [igtpos].[dbo].Family\n"
        "\tWHERE Name = @familia\n"
        ");\n"
        "SET @idCategoria = @idFamilia;\n\n"
    
        "INSERT INTO [igtpos].[dbo].[IndexedGroup]\n"
        "SELECT 'Category_' + CONVERT(NVARCHAR(MAX), @idCategoria)\n"
        "WHERE NOT EXISTS\n"
        "(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].[IndexedGroup]\n"
        "\tWHERE KeyName = 'Category_' + CONVERT(NVARCHAR(MAX), @idCategoria)\n"
        ");\n\n"
        
        "SET @idFamilia = (SELECT Id FROM [igtpos].[dbo].Family WHERE [Name] = '%s');\n"
        "SET @idCategoria = @idFamilia;\n\n"
        % (nombre, texto_boton, nombre, nombre, texto_boton, nombre, nombre)
    )

    glb.log_file.write("INSERTADA LA FAMILIA %s EN LA BASE DE DATOS\n" % nombre)

    f = cache_traducciones_size // 2
    x1 = 0
    x2 = cache_traducciones_size - 1

    encontrado = False

    while not encontrado:
        if traducciones["A" + str(f + 2)].value == nombre:
            encontrado = True
            break

        elif traducciones["A" + str(f + 2)].value < nombre:
            x1 = f
            f = (x1 + x2) // 2

        else:
            x2 = f
            f = (x1 + x2) // 2

        if abs(x2 - x1) <= 1:
            if traducciones["A" + str(x1 + 2)].value == nombre:
                encontrado = True
                f = x1 + 2

            if traducciones["A" + str(x2 + 2)].value == nombre:
                encontrado = True
                f = x2 + 2

            break

    f += 2

    for idioma in list(glb.idiomas.keys()):
        codigo = glb.codigos[idioma]

        # Primero comprobamos si existe una traducción del texto en la caché

        c = 0

        while traducciones[glb.columnas[c] + "1"].value != codigo:
            c += 1

        columna = glb.columnas[c]

        # Si se ha encontrado la traducción, se elegirá desde la cache.
        # Si no se ha encontrado la traducción, se guardará en la cache

        if encontrado:
            if traducciones[columna + str(f)].value is not None:
                traduccion = traducciones[columna + str(f)].value
            else:
                traduccion = ud(str(translator.translate_text(nombre, target_lang=glb.idiomas[idioma])))
                traducciones[columna + str(f)].value = traduccion
        else:
            f = cache_traducciones_size + 2
            traduccion = ud(str(translator.translate_text(nombre, target_lang=glb.idiomas[idioma])))
            traducciones["A" + str(f)].value = nombre
            traducciones[columna + str(f)].value = traduccion
            glb.log_file.write("ACTUALIZADA EN CACHE LA TRADUCCIÓN DE %s al valor %s\n" % (
                nombre, traduccion.replace("'", "''")))

        archivo_cache.save(glb.ruta_cache)
        traduccion = traduccion.replace("'", "''")

        query += (
            "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItem\n"
            "SELECT '%s', '%s', '%s'\n"
            "WHERE NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].DigitalMenuResourceItem\n"
            "\tWHERE LanguageCode = '%s'\n"
            "\tAND Text = '%s'\n"
            ");\n\n"
            % (codigo, nombre.replace("'", "''"), traduccion, codigo, nombre.replace("'", "''"))
        )

        query += (
            "SELECT @idTraduccion = Id\n"
            "FROM [igtpos].[dbo].DigitalMenuResourceItem\n"
            "WHERE LanguageCode = '%s'\n"
            "AND Text = '%s';\n\n"
            % (codigo, nombre.replace("'", "''"))
        )

        query += (
            "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItems\n"
            "SELECT 1, @idTraduccion\n"
            "WHERE NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].DigitalMenuResourceItems\n"
            "\tWHERE DigitalMenuResourceItemId = @idTraduccion\n"
            ");\n\n"
        )

    """
    Paso 4: Insertar la imagen en la tabla [igtpos].[dbo].Image, solo si esta existe
    """
    if os.path.exists(ruta_imagen):
        query += (
            "SET @SQL_QUERY = N'\n"
            "SELECT @contenidoImagen = BulkColumn\n"
            "FROM OPENROWSET(BULK ''%s'', SINGLE_BLOB) AS ImagenBinaria\n"
            "WHERE BulkColumn IS NOT NULL;';\n\n"
            
            "EXEC sp_executesql @SQL_QUERY, N'@contenidoImagen VARBINARY (MAX) OUTPUT', @contenido OUTPUT;\n\n"
            
            "INSERT INTO [igtpos].[dbo].Image\n"
            "SELECT NEWID(), @contenido;\n\n"
                    
            "UPDATE [igtpos].[dbo].Family\n"
            "SET StyleImageId =\n"
            "(\n"
            "\tSELECT MIN(Id)\n"
            "\tFROM [igtpos].[dbo].Image\n"
            "\tWHERE Content = @contenido\n"
            ")\n"
            "WHERE Id = @idFamilia AND DeletionDate IS NULL;\n\n"
                    
            "UPDATE [igtpos].[dbo].Category\n"
            "SET StyleImageId =\n"
            "(\n"
            "\tSELECT MIN(Id)\n"
            "\tFROM [igtpos].[dbo].Image\n"
            "\tWHERE Content = @contenido\n"
            ")\n"
            "WHERE Id = @idCategoria AND DeletionDate IS NULL;\n\n"
            % ruta_imagen
        )

        # Si la categoria no tiene productos vendibles como principal, no se muestra en POS

        vendible_principal = False

        i = 2
        while hoja_productos_salida["A" + str(i)].value is not None:
            if (
                    hoja_productos_salida["A" + str(i)].value == nombre and
                    hoja_productos_salida["P" + str(i)].value == "Sí"):
                vendible_principal = True
                break

            i += 1

        if not vendible_principal:
            query += (
                "UPDATE[igtpos].[dbo].Category\n"
                "SET ShowInPos = 0\n"
                "WHERE Id = @idCategoria;\n\n"
            )

        indice = -1

        if nombre in categorias_entrada:
            indice = categorias_entrada.index(nombre)
        else:
            indice = fuera_de_entrada
            fuera_de_entrada += 1

        # Establecer orden de la familia en POS
        query += (
            "IF NOT EXISTS (SELECT 1 FROM [igtpos].[dbo].[IndexItem] "
            "WHERE IndexedGroupId = 1 AND EntityTypeName = 'Category' AND EntityId = @idCategoria)\n"
            "BEGIN\n"
            "\tINSERT INTO [igtpos].[dbo].[IndexItem]\n"
            "\tVALUES (1, 'Category', @idCategoria, %d);\n"
            "END\n"
            "ELSE\n"
            "BEGIN\n"
            "\tUPDATE [igtpos].[dbo].[IndexItem]\n"
            "\tSET IndexedItemIndex = %d\n"
            "\tWHERE IndexedGroupId = 1\n"
            "\tAND EntityTypeName = 'Category'\n"
            "\tAND EntityId = @idCategoria;\n"
            "END\n\n"
            % (indice, indice)
        )

    sys.stdout.write("\r%d / %d categorías procesadas" % (a, limit))
    sys.stdout.flush()
    
    a += 1

print()
script_sql.write(query)
script_sql.close()
glb.log_file.close()
