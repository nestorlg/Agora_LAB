import openpyxl as op
import configparser as cp
import deepl as dl
import sys

from unidecode import unidecode as ud

import Global as glb

"""
Paso 1: Leer la hoja de Productos del Excel de Salida y obtener, por cada producto, su categoría, tipo de preparación,
orden de preparación (el cual indica si es un producto padre o no), descripción y lista de alérgenos
"""

script_sql = open("sql/products.sql", "w")

productos = []

with open("temp.txt", "r") as f:
    lineas = f.readlines()

project = lineas[0][:-1]
log_file_name = lineas[1]

glb.datos_import = "projects/" + project + "/entrada.xlsx"
glb.datos_export = "projects/" + project + "/salida.xlsx"
glb.dir_imagenes = "imagenesAgora"
glb.log_file = open("projects/" + project + "/log_files/%s" % log_file_name + ".log", "a")

glb.ruta_cache = "outputExcelCreator/data/cache.xlsx"

translator = dl.Translator(glb.auth_key)
archivo_cache = op.load_workbook(glb.ruta_cache)
traducciones = archivo_cache["Traducciones"]

salida = op.load_workbook(glb.datos_export)

hoja_productos_salida = salida["Productos"]

cache_traducciones_size = 0

while traducciones["A" + str(cache_traducciones_size + 2)].value is not None:
    cache_traducciones_size += 1

limit = 0

i = 2
while hoja_productos_salida["A" + str(i)].value is not None:
    nombre = hoja_productos_salida["E" + str(i)].value
    categoria = hoja_productos_salida["B" + str(i)].value
    tipo_de_preparacion = hoja_productos_salida["V" + str(i)].value
    orden_de_preparacion = hoja_productos_salida["W" + str(i)].value
    descripcion = hoja_productos_salida["Z" + str(i)].value
    lista_alergenos = hoja_productos_salida["C" + str(i)].value

    producto = [
        nombre,
        categoria,
        tipo_de_preparacion,
        orden_de_preparacion,
        descripcion,
        lista_alergenos
    ]

    if orden_de_preparacion is not None:
        productos.append(producto)
        limit += 1

    i += 1

productos_entrada = []

entrada = op.load_workbook(glb.datos_import)

hoja_productos_entrada = entrada["Productos"]

i = 2
while hoja_productos_entrada["A" + str(i)].value is not None:
    nombre_sin_formato = hoja_productos_entrada["A" + str(i)].value.split(" (")[0]
    categoria = hoja_productos_entrada["C" + str(i)].value
    if [nombre_sin_formato, categoria] not in productos_entrada:
        productos_entrada.append([nombre_sin_formato, categoria])

    i += 1

query = (
    "SET NOCOUNT ON;\n\n"
    
    "DECLARE @idTipoPreparacion INT;\n"
    "DECLARE @idOrdenPreparacion INT;\n"
    "DECLARE @idFamilia INT;\n"
    "DECLARE @idCategoria INT;\n"
    "DECLARE @idAlergeno INT;\n"
    "DECLARE @idProducto INT;\n"
    "DECLARE @idGrupo INT;\n"
    "DECLARE @SQL_QUERY NVARCHAR(MAX);\n"
    "DECLARE @contenido VARBINARY(MAX);\n"
    "DECLARE @idTraduccion INT;\n\n"
)

k = 1
for producto in productos:
    """
    Paso 2: Obtener el id del tipo de preparación, orden de preparación, familia y categoría
    """

    nombre = producto[0]
    categoria = producto[1]
    tipo_de_preparacion = producto[2]
    orden_de_preparacion = producto[3]
    descripcion = producto[4]
    alergenos = []
    if producto[5] is not None:
        alergenos = producto[5].split(", ")

    query += (
        "SELECT @idTipoPreparacion = Id\n"
        "FROM [igtpos].[dbo].PreparationType\n"
        "WHERE Name = '%s';\n\n"
        % tipo_de_preparacion
    )

    query += (
        "SELECT @idOrdenPreparacion = Id\n"
        "FROM [igtpos].[dbo].PreparationOrder\n"
        "WHERE Name = '%s';\n\n"
        % orden_de_preparacion
    )

    query += (
        "SELECT @idFamilia = Id\n"
        "FROM [igtpos].[dbo].Family\n"
        "WHERE Name = '%s';\n\n"
        % categoria
    )

    query += (
        "SELECT @idCategoria = Id\n"
        "FROM [igtpos].[dbo].Category\n"
        "WHERE Name = '%s';\n\n"
        % categoria
    )

    """
    Paso 3: Insertar el producto en la tabla [igtpos].[dbo].Product
    """

    query += (
        "INSERT INTO [igtpos].[dbo].Product\n"
        "SELECT 'P', NULL, '%s', '', 0, 0, 0, 1, 0, 0, 0, 0, 1, 0, NULL, NULL, @idTipoPreparacion, @idOrdenPreparacion,"
        " @idFamilia, 4, NULL, 0, '', 1, '', '%s', '', NULL, NULL, 0, 0, 0, 0, '', 1, NULL, NULL\n"
        "WHERE NOT EXISTS\n"
        "(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].Product\n"
        "\tWHERE [Name] = '%s'\n"
        ");\n\n"
        % (nombre.replace("'", "''"), categoria, nombre.replace("'", "''"))
    )

    glb.log_file.write("INSERTADO EL PRODUCTO %s EN LA BASE DE DATOS\n" % nombre)

    # Obtener el ID del producto que se acaba de insertar

    query += (
        "SELECT @idProducto = Id\n"
        "FROM [igtpos].[dbo].Product\n"
        "WHERE Name = '%s';\n\n"
        % nombre.replace("'", "''")
    )

    """
    Paso 4: Insertar los alérgenos del producto
    """

    for alergeno in alergenos:
        query += (
            "SELECT @idAlergeno =\n"
            "(\n"
            "\tSELECT Id\n"
            "\tFROM [igtpos].[dbo].Allergen\n"
            "\tWHERE Name = '%s'\n"
            ");\n\n"
            % alergeno.strip()
        )
        query += (
            "INSERT INTO [igtpos].[dbo].ProductAllergen\n"
            "SELECT @idProducto, @idAlergeno\n"
            "WHERE @idAlergeno != NULL AND NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].ProductAllergen\n"
            "\tWHERE ProductId = @idProducto\n"
            "\tAND AllergenId = @idAlergeno\n"
            ");\n\n"
        )

    """
    Paso 5: Relacionar el producto con la categoría
    """

    query += (
        "INSERT INTO [igtpos].[dbo].ProductCategory\n"
        "SELECT @idProducto, @idCategoria\n"
        "WHERE NOT EXISTS\n"
        "(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].ProductCategory\n"
        "\tWHERE CategoryId = @idCategoria\n"
        "\tAND ProductId = @idProducto\n"
        ");\n\n"
    )

    """
    Paso 6: Insertar la descripción del producto
    """

    if descripcion is not None:
        query += (
            "INSERT INTO [igtpos].[dbo].ProductDataSheet\n"
            "SELECT NULL, '%s', 0x0, @idProducto\n"
            "WHERE NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].ProductDataSheet\n"
            "\tWHERE ProductId = @idProducto\n"
            ");\n\n"
            % descripcion.replace("'", "''")
        )

        f = cache_traducciones_size // 2
        x1 = 0
        x2 = cache_traducciones_size - 1

        encontrado = False

        while not encontrado:
            if traducciones["A" + str(f + 2)].value == descripcion:
                encontrado = True
                break

            elif traducciones["A" + str(f + 2)].value < descripcion:
                x1 = f
                f = (x1 + x2) // 2

            else:
                x2 = f
                f = (x1 + x2) // 2

            if abs(x2 - x1) <= 1:
                if traducciones["A" + str(x1 + 2)].value == descripcion:
                    encontrado = True
                    f = x1 + 2

                if traducciones["A" + str(x2 + 2)].value == descripcion:
                    encontrado = True
                    f = x2 + 2

                break

        f += 2

        for idioma in list(glb.idiomas.keys()):
            codigo = glb.codigos[idioma]

            encontrado = False

            # Primero comprobamos si existe una traducción del texto en la caché

            c = 0

            while traducciones[glb.columnas[c] + "1"].value != codigo:
                c += 1

            columna = glb.columnas[c]

            f = 2
            while traducciones["A" + str(f)].value is not None:
                if traducciones["A" + str(f)].value == descripcion:
                    encontrado = True
                    break

                f += 1

            if encontrado:
                if traducciones[columna + str(f)].value is not None:
                    traduccion = traducciones[columna + str(f)].value
                else:
                    traduccion = ud(str(translator.translate_text(descripcion, target_lang=glb.idiomas[idioma])))
                    traduccion = traduccion.replace("'", "''")
                    traducciones[columna + str(f)].value = traduccion
            else:
                traduccion = ud(str(translator.translate_text(descripcion, target_lang=glb.idiomas[idioma])))
                traduccion = traduccion.replace("'", "''")
                traducciones["A" + str(f)].value = descripcion
                traducciones[columna + str(f)].value = traduccion
                glb.log_file.write("ACTUALIZADA EN CACHE LA TRADUCCIÓN DE %s al valor %s\n" % (
                    descripcion, traduccion.replace("'", "''")))

            archivo_cache.save(glb.ruta_cache)

            query += (
                "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItem\n"
                "SELECT '%s', '%s', '%s'\n"
                "WHERE NOT EXISTS\n"
                "(\n"
                "\tSELECT 1\n"
                "\tFROM [igtpos].[dbo].DigitalMenuResourceItem\n"
                "\tWHERE LanguageCode = '%s'\n"
                "\tAND Text = '%s'\n"
                ");\n\n"
                % (codigo, descripcion.replace("'", "''"), traduccion, codigo, descripcion.replace("'", "''"))
            )

            query += (
                "SELECT @idTraduccion = Id\n"
                "FROM [igtpos].[dbo].DigitalMenuResourceItem\n"
                "WHERE LanguageCode = '%s'\n"
                "AND Text = '%s';\n\n"
                % (codigo, descripcion.replace("'", "''"))
            )

            query += (
                "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItems\n"
                "SELECT 1, @idTraduccion\n"
                "WHERE NOT EXISTS\n"
                "(\n"
                "\tSELECT 1\n"
                "\tFROM [igtpos].[dbo].DigitalMenuResourceItems\n"
                "\tWHERE DigitalMenuResourceItemId = @idTraduccion\n"
                ");\n\n"
            )

    """
    Paso 7: Establecer el orden que tendrá el producto dentro de su categoría
    """

    i = 0
    ultima_categoria = ""

    for prod, cat in productos_entrada:
        if cat != ultima_categoria:
            i = 0

        ultima_categoria = cat
        if prod == nombre:
            break
        else:
            i += 1

    query += (
        "SELECT @idGrupo = Id\n"
        "FROM[igtpos].[dbo].IndexedGroup\n"
        "WHERE KeyName = 'Category_' + CONVERT(NVARCHAR(MAX), @idCategoria);\n\n"
    )

    query += (
        "INSERT INTO [igtpos].[dbo].IndexItem\n"
        "SELECT @idGrupo, 'Product', @idProducto, %d\n"
        "WHERE NOT EXISTS\n"
        "(\n"
        "\tSELECT 1\n"
        "\tFROM [igtpos].[dbo].IndexItem\n"
        "\tWHERE IndexedGroupId = @idGrupo\n"
        "\tAND EntityTypeName = 'Product'\n"
        "\tAND EntityId = @idProducto\n"
        "\tAND IndexedItemIndex = %d\n"
        ");\n\n"
        % (i, i)
    )

    sys.stdout.write("\r%d / %d productos base procesados" % (k, limit))
    sys.stdout.flush()

    k += 1

"""
Paso 8: Realizar la traducción de los alérgenos
"""

alergenos_a_traducir = [
    "Gluten",
    "Crustáceos",
    "Huevos",
    "Pescado",
    "Cacahuetes",
    "Soja",
    "Lácteos",
    "Frutos de cáscara",
    "Apio",
    "Mostaza",
    "Sésamo",
    "Sulfitos",
    "Altramuces",
    "Moluscos"
]

for alergeno_a_traducir in alergenos_a_traducir:
    f = cache_traducciones_size // 2
    x1 = 0
    x2 = cache_traducciones_size - 1

    encontrado = False

    while not encontrado:
        if traducciones["A" + str(f + 2)].value == alergeno_a_traducir:
            encontrado = True
            break

        elif traducciones["A" + str(f + 2)].value < alergeno_a_traducir:
            x1 = f
            f = (x1 + x2) // 2

        else:
            x2 = f
            f = (x1 + x2) // 2

        if abs(x2 - x1) <= 1:
            if traducciones["A" + str(x1 + 2)].value == alergeno_a_traducir:
                encontrado = True
                f = x1 + 2

            if traducciones["A" + str(x2 + 2)].value == alergeno_a_traducir:
                encontrado = True
                f = x2 + 2

            break

    f += 2

    for idioma in list(glb.idiomas.keys()):
        codigo = glb.codigos[idioma]

        # Primero comprobamos si existe una traducción del texto en la caché

        c = 0

        while traducciones[glb.columnas[c] + "1"].value != codigo:
            c += 1

        columna = glb.columnas[c]

        f = 2
        while traducciones["A" + str(f)].value is not None:
            if traducciones["A" + str(f)].value == alergeno_a_traducir:
                encontrado = True
                break

            f += 1

        if encontrado:
            if traducciones[columna + str(f)].value is not None:
                traduccion = traducciones[columna + str(f)].value
            else:
                traduccion = ud(str(translator.translate_text(alergeno_a_traducir, target_lang=glb.idiomas[idioma])))
                traducciones[columna + str(f)].value = traduccion
        else:
            f = cache_traducciones_size + 2
            traduccion = ud(str(translator.translate_text(alergeno_a_traducir, target_lang=glb.idiomas[idioma])))
            traducciones["A" + str(f)].value = alergeno_a_traducir
            traducciones[columna + str(f)].value = traduccion
            glb.log_file.write("ACTUALIZADA EN CACHE LA TRADUCCIÓN DE %s al valor %s\n" % (
                alergeno_a_traducir, traduccion.replace("'", "''")))

        archivo_cache.save(glb.ruta_cache)
        traduccion = traduccion.replace("'", "''")

        query += (
            "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItem\n"
            "SELECT '%s', '%s', '%s'\n"
            "WHERE NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].DigitalMenuResourceItem\n"
            "\tWHERE LanguageCode = '%s'\n"
            "\tAND Text = '%s'\n"
            ");\n\n"
            % (codigo, alergeno_a_traducir, traduccion, codigo, alergeno_a_traducir)
        )

        query += (
            "SELECT @idTraduccion = Id\n"
            "FROM [igtpos].[dbo].DigitalMenuResourceItem\n"
            "WHERE LanguageCode = '%s'\n"
            "AND Text = '%s';\n\n"
            % (codigo, alergeno_a_traducir)
        )

        query += (
            "INSERT INTO [igtpos].[dbo].DigitalMenuResourceItems\n"
            "SELECT 1, @idTraduccion\n"
            "WHERE NOT EXISTS\n"
            "(\n"
            "\tSELECT 1\n"
            "\tFROM [igtpos].[dbo].DigitalMenuResourceItems\n"
            "\tWHERE DigitalMenuResourceItemId = @idTraduccion\n"
            ");\n\n"
        )

print()
script_sql.write(query)
script_sql.close()
glb.log_file.close()
